from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.common.exceptions import WebDriverException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import platform
from bs4 import BeautifulSoup
import re
import openpyxl
from openpyxl import load_workbook
import random
import time
import os
import logging


# Configure logging at the beginning of the script
logging.basicConfig(filename='igScraper.log', level=logging.INFO, 
    format='%(asctime)s:%(levelname)s:%(message)s', filemode='a')

# Access the environment variables. Set these in terminal as either temporary or permanent.
# Must run program through terminal environemnt where these variables were set.
USERNAME = os.getenv('INSTAGRAM_USERNAME')
PASSWORD = os.getenv('INSTAGRAM_PASSWORD')

# Log the values to verify they are correctly retrieved
logging.info(f"USERNAME: {USERNAME}")
logging.info(f"PASSWORD: {PASSWORD}")

# Configurable parameters. Only have to change these here.
PROFILE_NAME = 'tradedny'
NUM_POSTS_TO_SCRAPE = 3000
OUTPUT_FILE_NAME = 'instagram202_data.xlsx'

# The ChromeDriver path is specific to your system.
# 'chromedriver' is a driver that Selenium uses to open up a new Google Chrome browser.

# chrome_driver_path = '/Users/jonathanbachrach/Documents/Automate/TradedScrape/chromedriver'

# ChromeDriver path
chrome_driver_path = '/usr/local/bin/chromedriver'

# 'Options' is a class in Selenium allowing us to customize & add arguments to Chrome browser.
# The options added here are to prevent any GUI from showing up. 
chrome_options = Options()
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.binary_location = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"

# Function to extract hashtags
def extract_hashtags(content):
    # the findall() function from the 're' module returns all non-overlapping matches 
    # of pattern in string, as a list of strings.
    hashtags = re.findall(r"#(\w+)", content)
    return ' '.join(['#' + tag for tag in hashtags])

# Function to parse content from each Instagram post.
def parse_content(content):
    # Parse the content with BeautifulSoup
    soup = BeautifulSoup(content, 'html.parser')
    h1 = soup.find('h1', class_='_aacl _aaco _aacu _aacx _aad7 _aade')
    if h1:
        first_word = h1.get_text(separator=" ").split()[0]
    else:
        first_word = None

    # Define an empty string for hashtags.
    hashtags_str = extract_hashtags(content)

    # Extract first word of the h1 tag, and the hashtags, and create a dictionary with these.
    data = {'tradedny': first_word, 'hashtags': hashtags_str}
    
    # Split the content into lines (it's currently one long string).
    lines = content.split("\n")

    label_dict = {"BROKERS": "BROKER", "NOTE FROM BROKER": "NOTE", "BUYERS": "BUYER", "BUYER'S": "BUYER",
              "SELLERS": "SELLER", "SELLER'S": "SELLER", "BUYERS REP": "BUYER'S REP", "SELLERS REP": "SELLER'S REP",
              "TENANT'S REP": "TENANT REP", "UNIT": "UNITS"}


    # Iterate over each line.
    for line in lines:
        # Split each line into parts separated by "~".
        parts = line.strip().split("~")

        # Process each part seperately.
        for part in parts:
            # Strip leadind/trailing spaces and split part into label and value,
            #  only at the first occurrance of ":"
            sub_parts = part.strip().split(": ", 1)

            # Make sure we have both a label and a value.
            if len(sub_parts) == 2:
                label = sub_parts[0].strip().upper() # Normalizing labels to match header names
                value = sub_parts[1].strip()

                # Use the dictionary to normalize the label. If the label is not
                # in the dictionary, use the label as is.
                label = label_dict.get(label, label)


                # Clean up the value by parsing it with BeautifulSoup and extracting
                # the text.
                value_soup = BeautifulSoup(value, 'html.parser')
                value = value_soup.get_text()

                data[label] = value

    return data

# Function to handle the login.
def login(browser, username, password):
    browser.get('https://www.instagram.com')

    try:
        userElem = WebDriverWait(browser, 25).until(EC.presence_of_element_located((By.NAME, 
            'username'))
        )
        userElem.send_keys(USERNAME)

        passwordElem = WebDriverWait(browser, 25).until(EC.presence_of_element_located((By.NAME,
        'password')))
        passwordElem.send_keys(PASSWORD)

        passwordElem.send_keys(Keys.RETURN)

        time.sleep(25)

        notification_button = WebDriverWait(browser, 25).until(EC.presence_of_element_located((By.CLASS_NAME, '_a9--')))
        notification_button.click()

        time.sleep(15)
    except (TimeoutException, NoSuchElementException) as e:
        logging.error(f"Login failed: {e}")
        logging.error("Pausing for 1 minute before retrying login...")
        time.sleep(60)
        login(browser, username, password)  # Retry the login

# Function to navigate to the profile.
def load_profile(browser, profile_name):
    try:
        """Go to tradedny instagram page."""
        search_element = browser.find_element(By.CSS_SELECTOR, 
            '.x1xgvd2v > div:nth-child(2) > div:nth-child(2) > '
            'span:nth-child(1) > div:nth-child(1) > a:nth-child(1) > '
            'div:nth-child(1) > div:nth-child(2) > div:nth-child(1) > '
            'div:nth-child(1) > span:nth-child(1) > span:nth-child(1)')

        browser.execute_script("arguments[0].scrollIntoView();", search_element)
        time.sleep(2)

        # Click the search element
        browser.execute_script("arguments[0].click();", search_element)

        time.sleep(5)

        # Wait for the search input field and enter the profile name

        type_search = WebDriverWait(browser, 20).until( 
            EC.presence_of_element_located((By.CSS_SELECTOR, 'input.x1lugfcp'))
        )

        type_search.send_keys(PROFILE_NAME)

        # Wait for results to load
        time.sleep(15)

        # Locate the profile link element

        type_click = browser.find_element(
            By.XPATH, 
            "//span[contains(@class, 'x1lliihq') and contains(text(), 'Traded: New York🗽')]" 
        )
    
        # Scroll the profile element into view    
        browser.execute_script("arguments[0].scrollIntoView();", type_click)
        time.sleep(2)

        # Click on the profile link
        type_click.click()
        # Allow time for the page to navigate
        time.sleep(20)
    except (TimeoutException, NoSuchElementException) as e:
        logging.error(f"Loading profile failed: {e}")
        logging.error("Pausing for 5 minutes before retrying...")
        time.sleep(300)
        load_profile(browser, profile_name)  # Retry loading the profile


# Function to scroll down the page to load all posts.
def scroll_down(browser):
    """Scroll to the end of the page to load all posts (add as needed).
    scrollIntoView() method scrolls the specified element into the visible area 
    of the browser window."""

    browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(15)

# Function to click on first post.
def click_on_post(browser): 
    try:
        first_post = WebDriverWait(browser, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 
            "img.x5yr21d.xu96u03.x10l6tqk.x13vifvy.x87ps6o.xh8yej3"))
        )

        # Click the element via JavaScript
        browser.execute_script("arguments[0].click();", first_post)
        time.sleep(4)
    except (TimeoutException, NoSuchElementException) as e:
        logging.error(f"Clicking on the first post failed: {e}")
        logging.error("Pausing for 3 minutes before retrying...")
        time.sleep(180)
        click_on_post(browser)  # Retry clicking the post


# Function to parse a post.
def parse_post(browser):
    try:
        # Extract the post.
        content = WebDriverWait(browser, 25).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div[class*='_a9zs']"))
            ).get_attribute('innerHTML')
        content = content.replace('<br>', '\n')

        # Parse the content.
        return parse_content(content)

    except Exception as e:
        logging.error(f"Unexpected error in parse_post: {e}")
        return None 


# Function to add parsed data to the Excel spreadsheet.
def save_data(browser, data, headers, sheet):
    """The data is a list of values, with each value 
    corresponding to a header in the Excel spreadsheet."""
    # If a header does not have a corresponding value in the data, 
    # use a space " " as a placeholder.
    row = [data.get(header, " ") for header in headers]
    sheet.append(row)


def start_browser(): 
    # Setup the ChromeDriver service with the specified path
    try: 

        service = Service(chrome_driver_path)

        # Set up Chrome options
        options = Options()
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.binary_location = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"

        logging.info('Initializing ChromeDriver...')

        # Initialize the WebDriver with the correct service and options
        browser = webdriver.Chrome(service=service, options=options)

        logging.info('ChromeDriver initialized. Starting Chrome...')

        logging.info('Chrome started successfully.')
    except Exception as e: 
        logging.error(f"Failed to initialize ChromeDriver: {e}")
        print(f"Failed to initialize ChromeDriver: {e}")
        exit(1)

    return browser


def read_last_scraped_index():
    """ Check if the file exists and read the last scraped index."""
    if os.path.exists("last_scraped_index.txt"):
        with open("last_scraped_index.txt", "r") as file:
            return int(file.read().strip())
    else:
        return 0


def write_last_scraped_index(index):
    # Write the last scraped index to a file
    with open("last_scraped_index.txt", "w") as file:
        file.write(str(index))


def initialize_browser_and_workbook():
    """Initialize the browser and load or create the Excel workbook."""
    browser = start_browser()
    time.sleep(5)
    browser.get("https://www.instagram.com")
    logging.info('Starting the scraping process...') 

    # Define headers
    headers = ["tradedny", "IMAGE", "DATE", "ADDRESS", "MARKET", "ASSET TYPE", "LENDER", "BUYER",
        "RENTER", "SELLER", "LANDLORD", "SELLER'S REP", "BUYER'S REP", "LOAN AMOUNT", "LOAN TYPE", "TENANT", "TENANT REP",
        "LANDLORD REP", "BROKER", "SALE PRICE", "ASKING RENT", "SF", "PPSF", "UNITS",
        "PPU", "BSF", "PPBSF", "NOTE", "hashtags"]

    # Load existing workbook if it already exists
    if os.path.exists(OUTPUT_FILE_NAME):
        logging.info('Loading existing Excel workbook...')
        wb = load_workbook(OUTPUT_FILE_NAME)
        sheet = wb.active

    else: 
        # Create a new workbook if one doesn't exist
        logging.info('Initializing the Excel workbook...')
        wb = openpyxl.Workbook()
        sheet = wb.active
        # Insert the headers to the first row.
        sheet.append(headers)

    # Debugging: Log the outputs
    if browser and wb and sheet and headers: 
        logging.info('Browser, workbook, sheet, and headers initialized successfully.')
    else: 
        logging.error('One or more of the initialization components is None.')

    return browser, wb, sheet, headers 


def scrape_instagram(browser, wb, sheet, last_scraped_index, headers):
    """Perform the Instagram scraping process."""

    try: 

        logging.info('Attempting to log in...')
        login(browser, USERNAME, PASSWORD)
        logging.info('Login successful, proceeding to load profile.')
    except Exception as e: 
        logging.error(f"Login failed: {e}")
        return  # Exit the function if login

    # Load profile and start scraping
    load_profile(browser, PROFILE_NAME)
    scroll_down(browser)
    click_on_post(browser)

    # Counter for consecutive errors
    consecutive_errors = 0
    max_consecutive_errors = 3

    # Begin the loop to click on posts and parse them.
    for i in range(last_scraped_index, NUM_POSTS_TO_SCRAPE):
        logging.info(f'Scraping post {i+1} of {NUM_POSTS_TO_SCRAPE}...')
        try:
            # Add a random delay before each action
            time.sleep(random.uniform(3, 8))

            # Extract the content
            logging.info('Parsing the post content...')
            data = parse_post(browser)

            # If data is None, this means an error occurred while parsing the post. 
            if data is None:
                logging.warning(f"No data returned for post {i+1}.")
                consecutive_errors += 1
            else:
                # If data is not None, the post was successfully parsed. Reset error counter.
                logging.info(f"Data extracted for post {i+1}: {data}")
                consecutive_errors = 0

                # Save the data.
                logging.info('Saving post data to Excel...')
                save_data(browser, data, headers, sheet)
                # Update the last scraped index
                write_last_scraped_index(i)

                # Save every 10 posts
                if i % 10 == 0:
                    logging.info('Saving the Excel workbook...')
                    # Save the workbook after each post is processed
                    wb.save(OUTPUT_FILE_NAME)

            # If there were more than 3 consecutive errors, skip the post
            if consecutive_errors > max_consecutive_errors:
                logging.error(f"Too many errors encountered. Stopping the program.")
                raise Exception("Too many consecutive errors. Exiting the program.")

            # Add a random delay.
            time.sleep(random.uniform(3, 8))

            # Click on the 'Next' button to go to the next post.
            next_button = WebDriverWait(browser, 20).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "svg[aria-label='Next']"))
            )
                
            next_button.click()

            # Add a random delay.
            time.sleep(random.uniform(3, 8))

        except (WebDriverException, TimeoutException) as e:
            logging.error(f"Error: {e}")
            consecutive_errors += 1

            if "disconnected" in str(e) or consecutive_errors >= max_consecutive_errors:
                raise

            # Continue to the next post after logging the error
            continue

        except Exception as e: 
            # Log any other exceptions and stop the program
            logging.error(f"Unexpected error: {e}")
            raise  # Re-raise the exception to display the traceback and stop the program


def save_workbook_and_cleanup(wb, browser):
    """Save the workbook and clean up resources."""
    if wb: 
        logging.info('Saving the Excel workbook...')
        wb.save(OUTPUT_FILE_NAME)

    if browser: 
        logging.info('Closing the browser...')
        browser.quit()

    if os.path.exists("last_scraped_index.txt"):
        os.remove("last_scraped_index.txt")
        logging.info('Deleted last_scraped_index.txt to start fresh next time.')


# Main function
def main():
    logging.info('Starting the scraping process...')

    # Initialize browser variable
    browser = None 
    # Read where we left off
    wb = None
    last_scraped_index = read_last_scraped_index() 

    try:
        browser, wb, sheet, headers = initialize_browser_and_workbook()
        if not all([browser, wb, sheet, headers]):
            logging.error("Initialization failed: One or more components are None.")
            return 

        while True:
            try: 
                # Perform the scraping process
                scrape_instagram(browser, wb, sheet, last_scraped_index, headers)
                # Exit the loop if scraping completes successfully
                break

            except WebDriverException as e:
                logging.error(f"WebDriverException: {e}")
                if "disconnected" in str(e):
                    logging.error("Chrome disconnected. Restarting the browser...")
                    if browser: 
                        browset.quit()
                    # Slight delay before attempting to restart
                    browser, wb, sheet, headers = initialize_browser_and_workbook()
                    if not all([browser, wb, sheet, headers]):
                        logging.error("Reinitialization failed after disconnection.")
                else: 
                    raise 

            except Exception as e:
                logging.error(f"Unexpected error in main loop: {e}")
                # Exit the loop on other unexpected errors
                break 


    finally: 
        save_workbook_and_cleanup(wb, browser)


# Call the main function.
if __name__ == "__main__":
    main()












    






    def scrape_instagram(browser, sheet, headers, num_posts=10):
        logging.info(f"Starting the scraping process...")

        profile_url = "https://www.instagram.com/tradedny/"
        browser.get(profile_url)
        time.sleep(5)

        # Click on the first post
        try:
            first_post = WebDriverWait(browser, 15).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "article a"))
            )
            first_post.click()
            time.sleep(5)
        except Exception as e:
            logging.error(f"Failed to open the first post: {e}")
            return

        consecutive_errors = 0

        for i in range(num_posts):
            logging.info(f"Scraping post {i + 1} of {num_posts}...")
            try:
                data = parse_post(browser)
                if data:
                    logging.info(f"Data extracted for post {i + 1}: {data}")
                    logging.info("Saving post data to Excel...")
                    row = [data.get(header, "") for header in headers]
                    sheet.append(row)
                else:
                    logging.warning(f"No data returned for post {i + 1}.")
                    consecutive_errors += 1
                    if consecutive_errors >= 3:
                        raise Exception("Too many consecutive errors. Exiting the program.")
                    continue

                consecutive_errors = 0

                if (i + 1) % 10 == 0:
                    logging.info("Saved Excel workbook after 10 posts.")

                # Click the "Next" arrow to go to the next post
                next_button = WebDriverWait(browser, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "div._aaqg button._abl-"))
                )
                next_button.click()
                time.sleep(5)
            except Exception as e:
                logging.error(f"Unexpected error scraping post {i + 1}: {e}", exc_info=True)
                consecutive_errors += 1
                if consecutive_errors >= 3:
                    logging.error("Too many consecutive errors scraping.")
                    raise Exception("Too many consecutive errors. Exiting the program.")
                continue

                # Function to parse a post.
                def parse_post(browser, last_alt_text):
                    logging.info("Parsing the post content...")

                    try:
                        # Try to get the <meta name="description"> content
                        try:

                            description_element = WebDriverWait(browser, 25).until(
                                EC.presence_of_element_located((By.XPATH, "//meta[@name='description']")))

                            # Extract the alt attribute where Instagram stores post metadata
                            new_alt_text = description_element.get_attribute("content")

                        except TimeoutException:
                            logging.warning("Meta description not found. Trying img[alt] fallback.")

                            # Fallback to img[alt] if meta tag not found
                            image_element = WebDriverWait(browser, 25).until(
                                EC.presence_of_element_located((By.CSS_SELECTOR, "img[alt]")))

                            new_alt_text = image_element.get_attribute("alt")

                        # Wait until the new alt/description is different from the last one
                        WebDriverWait(browser, 25).until(
                            lambda driver: new_alt_text and new_alt_text != last_alt_text)

                        if not new_alt_text:
                            logging.warning("Post has no alt or description text.")
                            return None, last_alt_text

                        logging.info("Post content successfully extracted from meta tag or alt attribute.")

                        # Parse the raw post text into structured data
                        lines = new_alt_text.split("\n")
                        post_data = {"tradedny": None, "hashtags": ""}

                        for line in lines:
                            if ":" in line:
                                key, value = line.split(":", 1)
                                post_data[key.strip().upper()] = value.strip()
                            elif line.startswith("#"):
                                post_data["hashtags"] += line.strip() + " "

                        logging.info(f"Data extracted for post: {post_data}")
                        return post_data, new_alt_text


                    except TimeoutException:
                        logging.error("Timeout while waiting for post to load or alt to change.")
                        return None, last_alt_text


                    except Exception as e:
                        logging.error(f"Unexpected error in parse_post: {type(e).__name__}: {e}")
                        logging.exception("Full traceback:")
                        return None, last_alt_text

                    except Exception as e:
                        logging.error(f"Unexpected error in parse_post: {type(e).__name__}: {e}")
                        logging.exception("Full traceback:")
                        return None, last_alt_text


# Function to parse content from each Instagram post.
def parse_content(content):
    # Parse the content with BeautifulSoup
    soup = BeautifulSoup(content, 'html.parser')
    h1 = soup.find('h1', class_='_aacl _aaco _aacu _aacx _aad7 _aade')
    if h1:
        first_word = h1.get_text(separator=" ").split()[0]
    else:
        first_word = None

    # Define an empty string for hashtags.
    hashtags_str = extract_hashtags(content)

    # Extract first word of the h1 tag, and the hashtags, and create a dictionary with these.
    data = {'tradedny': first_word, 'hashtags': hashtags_str}
    
    # Split the content into lines (it's currently one long string).
    lines = content.split("\n")

    label_dict = {"BROKERS": "BROKER", "NOTE FROM BROKER": "NOTE", "BUYERS": "BUYER", "BUYER'S": "BUYER",
              "SELLERS": "SELLER", "SELLER'S": "SELLER", "BUYERS REP": "BUYER'S REP", "SELLERS REP": "SELLER'S REP",
              "TENANT'S REP": "TENANT REP", "UNIT": "UNITS"}


    # Iterate over each line.
    for line in lines:
        # Split each line into parts separated by "~".
        parts = line.strip().split("~")

        # Process each part seperately.
        for part in parts:
            # Strip leadind/trailing spaces and split part into label and value,
            #  only at the first occurrance of ":"
            sub_parts = part.strip().split(": ", 1)

            # Make sure we have both a label and a value.
            if len(sub_parts) == 2:
                label = sub_parts[0].strip().upper() # Normalizing labels to match header names
                value = sub_parts[1].strip()

                # Use the dictionary to normalize the label. If the label is not
                # in the dictionary, use the label as is.
                normalized_label = label_dict.get(label, label).lower()

                # Clean up the value by parsing it with BeautifulSoup and extracting
                # the text.
                value_soup = BeautifulSoup(value, 'html.parser')
                value = value_soup.get_text()

                data[normalized_label] = value




