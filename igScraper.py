
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.common.exceptions import WebDriverException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import re
import openpyxl
from openpyxl import load_workbook
import random
import time
import os
import logging

# Access the environment variables. Set these in terminal as either temporary or permanent.
# Must run program through terminal environemnt where these variables were set.
USERNAME = os.getenv('INSTAGRAM_USERNAME')
PASSWORD = os.getenv('INSTAGRAM_PASSWORD')

# Configurable parameters. Only have to change these here.
PROFILE_NAME = 'tradedny'
NUM_POSTS_TO_SCRAPE = 4000
OUTPUT_FILE_NAME = 'instagram2024_data.xlsx'

# The ChromeDriver path is specific to your system.
# 'chromedriver' is a driver that Selenium uses to open up a new Google Chrome browser.
chrome_driver_path = '/Users/jonathanbachrach/Documents/Automate/TradedScrape/chromedriver'

# 'Options' is a class in Selenium allowing us to customize & add arguments to Chrome browser.
# The options added here are to prevent any GUI from showing up. 
chrome_options = Options()
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.binary_location = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"

# Function to extract hashtags
def extract_hashtags(content):
    # the findall() function from the 're' module returns all non-overlapping matches 
    # of pattern in string, as a list of strings.
    hashtags = re.findall(r"#(\w+)", content)
    return ' '.join(['#' + tag for tag in hashtags])

# Function to parse content from each Instagram post.
def parse_content(content):
    # Parse the content with BeautifulSoup
    soup = BeautifulSoup(content, 'html.parser')
    h1 = soup.find('h1', class_='_aacl _aaco _aacu _aacx _aad7 _aade')
    if h1:
        first_word = h1.get_text(separator=" ").split()[0]
    else:
        first_word = None

    # Define an empty string for hashtags.
    hashtags_str = extract_hashtags(content)

    # Extract first word of the h1 tag, and the hashtags, and create a dictionary with these.
    data = {'tradedny': first_word, 'hashtags': hashtags_str}
    
    # Split the content into lines (it's currently one long string).
    lines = content.split("\n")

    label_dict = {"BROKERS": "BROKER", "NOTE FROM BROKER": "NOTE", "BUYERS": "BUYER", "BUYER'S": "BUYER",
              "SELLERS": "SELLER", "SELLER'S": "SELLER", "BUYERS REP": "BUYER'S REP", "SELLERS REP": "SELLER'S REP",
              "TENANT'S REP": "TENANT REP", "UNIT": "UNITS"}


    # Iterate over each line.
    for line in lines:
        # Split each line into parts separated by "~".
        parts = line.strip().split("~")

        # Process each part seperately.
        for part in parts:
            # Strip leadind/trailing spaces and split part into label and value,
            #  only at the first occurrance of ":"
            sub_parts = part.strip().split(": ", 1)

            # Make sure we have both a label and a value.
            if len(sub_parts) == 2:
                label = sub_parts[0].strip().upper() # Normalizing labels to match header names
                value = sub_parts[1].strip()

                # Use the dictionary to normalize the label. If the label is not
                # in the dictionary, use the label as is.
                label = label_dict.get(label, label)


                # Clean up the value by parsing it with BeautifulSoup and extracting
                # the text.
                value_soup = BeautifulSoup(value, 'html.parser')
                value = value_soup.get_text()

                data[label] = value

    return data

# Function to handle the login.
def login(browser, username, password):
    browser.get('https://www.instagram.com')

    # use WebDriverWait Selenium func to pause the script until the 'username' and 'password' 
    # elements have been loaded into the DOM (Document Object Model).

    userElem = WebDriverWait(browser, 25).until(EC.presence_of_element_located((By.NAME,
     'username')))
    userElem.send_keys(USERNAME)

    passwordElem = WebDriverWait(browser, 25).until(EC.presence_of_element_located((By.NAME,
     'password')))
    passwordElem.send_keys(PASSWORD)

    passwordElem.send_keys(Keys.RETURN)

    time.sleep(25)

    # save_login_info = WebDriverWait(browser, 20).until(EC.presence_of_element_located((By.XPATH, 
    #     "//button[text()='Save Info']")))
    # save_login_info.click()

    notification_button = WebDriverWait(browser, 25).until(EC.presence_of_element_located((By.CLASS_NAME, '_a9--')))
    notification_button.click()

    time.sleep(15)


# Function to navigate to the profile.
def load_profile(browser, profile_name):
    """Go to tradedny instagram page."""
    search_element = browser.find_element(By.CSS_SELECTOR, 
        '.x1xgvd2v > div:nth-child(2) > div:nth-child(2) > '
        'span:nth-child(1) > div:nth-child(1) > a:nth-child(1) > '
        'div:nth-child(1) > div:nth-child(2) > div:nth-child(1) > '
        'div:nth-child(1) > span:nth-child(1) > span:nth-child(1)')

    browser.execute_script("arguments[0].scrollIntoView();", search_element)
    time.sleep(2)

    # Click the search element
    browser.execute_script("arguments[0].click();", search_element)

    time.sleep(5)

    # Wait for the search input field and enter the profile name
    type_search = WebDriverWait(browser, 20).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, 'input.x1lugfcp'))
    )
  
    type_search.send_keys(PROFILE_NAME)

    # Wait for results to load
    time.sleep(15)

    # Locate the profile link element

    type_click = browser.find_element(By.XPATH, "//span[contains(@class, 'x1lliihq') and contains(text(), 'Traded: New York🗽')]")
    
    # Scroll the profile element into view    
    browser.execute_script("arguments[0].scrollIntoView();", type_click)
    time.sleep(2)

    # Click on the profile link
    type_click.click()

    # Allow time for the page to navigate
    time.sleep(20)


# Function to scroll down the page to load all posts.
def scroll_down(browser):
    """Scroll to the end of the page to load all posts (add as needed).
    scrollIntoView() method scrolls the specified element into the visible area 
    of the browser window."""

    browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(15)

# Function to click on first post.
def click_on_post(browser):
    first_post = WebDriverWait(browser, 20).until(EC.presence_of_element_located((By.CSS_SELECTOR, 
    "img.x5yr21d.xu96u03.x10l6tqk.x13vifvy.x87ps6o.xh8yej3")))

    # Click the element via JavaScript
    browser.execute_script("arguments[0].click();", first_post)
    time.sleep(4)


# Function to parse a post.
def parse_post(browser):
    # If error encountered during parsing, function returns a value of None.
    data = None
    try:
        # Extract the post.
        content = WebDriverWait(browser, 25).until(EC.presence_of_element_located((By.CSS_SELECTOR, 
            "div[class*='_a9zs']"))).get_attribute('innerHTML')

        # Replace <br> with newline character \n.
        content = content.replace('<br>', '\n')

        # Parse the content.
        data = parse_content(content)

    except NoSuchElementException as e:
        print(f"Error: {e}")
        print("Could not find the element. Skipping to the next post...")

    except TimeoutException as e:
        print(f"Error: {e}")
        print("Timed out waiting for page to load. Skipping to the next post...")

    except Exception as e:
        print(f"Unexpected error: {e}")

    return data


# Function to add parsed data to the Excel spreadsheet.
def save_data(browser, data, headers, sheet):
    """The data is a list of values, with each value 
    corresponding to a header in the Excel spreadsheet."""
    # If a header does not have a corresponding value in the data, 
    # use a space " " as a placeholder.
    row = [data.get(header, " ") for header in headers]
    sheet.append(row)


def start_browser(): 
    # Start and return a new browser session
    options = webdriver.ChromeOptions()
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.binary_location = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"

    logging.info('Initializing ChromeDriver...')

    # Install the specific versionof ChromeDriver
    service = Service(ChromeDriverManager().install())

    logging.info('ChromeDriver initialized. Starting Chrome...')

    browser = webdriver.Chrome(service=service, options=options)

    logging.info('Chrome started successfully.')

    return browser


def read_last_scraped_index():
    """ Check if the file exists and read the last scraped index."""
    if os.path.exists("last_scraped_index.txt"):
        with open("last_scraped_index.txt", "r") as file:
            return int(file.read().strip())
    else:
        return 0


def write_last_scraped_index(index):
    # Write the last scraped index to a file
    with open("last_scraped_index.txt", "w") as file:
        file.write(str(index))


# Main function
def main():
    # Initialize browser variable
    browser = None 
    # Read where we left off
    last_scraped_index = read_last_scraped_index()

    try:
        browser = start_browser()
        browser.get("https://www.instagram.com")
        # Set up logging
        # The log file 'igScraper.log' will contain all logs.
        logging.basicConfig(filename='igScraper.log', level=logging.INFO,
            format='%(asctime)s:%(levelname)s:%(message)s', filemode='a')
        logging.info('Starting the scraping process...')

        # Load existing workbook if it already exists
        if os.path.exists(OUTPUT_FILE_NAME):
            logging.info('Loading existing Excel workbook...')
            wb = load_workbook(OUTPUT_FILE_NAME)
            sheet = wb.active

        # Create a new workbook if one doesn't exist
        else: 
            logging.info('Initializing the Excel workbook...')
            wb = openpyxl.Workbook()
            sheet = wb.active

            # Insert the headers to the first row.
            headers = ["tradedny", "IMAGE", "DATE", "ADDRESS", "MARKET", "ASSET TYPE", "LENDER", "BUYER",
                "RENTER", "SELLER", "LANDLORD", "SELLER'S REP", "BUYER'S REP", "LOAN AMOUNT", "LOAN TYPE", "TENANT", "TENANT REP",
                "LANDLORD REP", "BROKER", "SALE PRICE", "ASKING RENT", "SF", "PPSF", "UNITS",
                "PPU", "BSF", "PPBSF", "NOTE", "hashtags"]

            sheet.append(headers)

        # Login
        logging.info('Logging in to Instagram')
        login(browser, USERNAME, PASSWORD)

        # Load the profile.
        logging.info('Loading Instagram profile...')
        load_profile(browser, PROFILE_NAME)

        # Scroll down to load all posts.
        logging.info('Scrolling down to load all posts...')
        scroll_down(browser)

        # Click on first post.
        logging.info('Clicking on the first post...')
        click_on_post(browser)

        # Counter for consecutive errors
        consecutive_errors = 0

        # Begin the loop to click on posts and parse them.
        for i in range(last_scraped_index, NUM_POSTS_TO_SCRAPE):
            logging.info(f'Scraping post {i+1} of {NUM_POSTS_TO_SCRAPE}...')
            try:
                # Add a random delay before each action
                time.sleep(random.uniform(3, 8))

                # Extract the content
                logging.info('Parsing the post content...')
                data = parse_post(browser)

                # If data is None, this means an error occurred while parsing the post. 
                if data is None:
                    consecutive_errors += 1
                else:
                    # If data is not None, the post was successfully parsed. Reset error counter.
                    consecutive_errors = 0

                    # Save the data.
                    logging.info('Saving post data to Excel...')
                    save_data(browser, data, headers, sheet)
                    # Update the last scraped index
                    write_last_scraped_index(i)

                # If there were more than 3 consecutive errors, break the loop.
                if consecutive_errors > 3:
                    logging.error("More than 3 consecutive errors in post. Breaking the loop...")
                    break

                # Add a random delay.
                time.sleep(random.uniform(3, 8))

                # Click on the 'Next' button to go to the next post.
                logging.info('Navigating to the next post...')
                next_button = browser.find_element(By.CSS_SELECTOR, "svg[aria-label='Next']")
                next_button.click()

                # Add a random delay.
                time.sleep(random.uniform(3, 8))

            # Specific exception handling
            except NoSuchElementException as e:
                logging.error(f"Error: {e}")
                logging.error("Could not find the element. Skipping to the next post...")
                continue

            except TimeoutException as e:
                logging.error(f"Error: {e}")
                logging.error("Timed out waiting for page to load. Skipping to the next post...")
                continue

            except WebDriverException as e: 
                logging.error(f"WebDriverException: {e}")
                if "disconnected" in str(e): 
                    logging.error("Chrome disconnected. Restarting the browser...")
                    browser.quit()
                    browser = start_browser()
                    login(browser, USERNAME, PASSWORD)
                    load_profile(browser, PROFILE_NAME)
                    scroll_down(browser)
                    click_on_post(browser)
                else:
                    raise


            except Exception as e:
                logging.error(f"Unexpected error: {e}")
                continue

        # Save the workbook then close the browser.
        logging.info('Saving the Excel workbook...')
        wb.save(OUTPUT_FILE_NAME)

    except Exception as e:
        logging.error(f"Unexpected error in main loop: {e}")

    finally:
        # Ensure browser is not None before quitting
        if browser: 
            logging.info('Closing the browser...')
            browser.quit()

    logging.info('Finished Instagram scraping process.')


# Call the main function.
if __name__ == "__main__":
    main()

    
    