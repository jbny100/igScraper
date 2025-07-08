
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.common.exceptions import WebDriverException, InvalidSessionIdException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
import platform
from bs4 import BeautifulSoup
import re
import openpyxl
from openpyxl import load_workbook
import random
import time
import os
import logging


# Configure logging at the beginning of the script
logging.basicConfig(filename='igScraper.log', level=logging.INFO, 
    format='%(asctime)s:%(levelname)s:%(message)s', filemode='a')

# Access the environment variables. Set these in terminal as either temporary or permanent.
# Must run program through terminal environemnt where these variables were set.
USERNAME = os.getenv('INSTAGRAM_USERNAME')
PASSWORD = os.getenv('INSTAGRAM_PASSWORD')

# Log the values to verify they are correctly retrieved
logging.info(f"USERNAME: {USERNAME}")
logging.info(f"PASSWORD: {PASSWORD}")

# Configurable parameters. Only have to change these here.
PROFILE_NAME = 'tradedny'
NUM_POSTS_TO_SCRAPE = 3500
OUTPUT_FILE_NAME = 'instagram2025_data.xlsx'

# The ChromeDriver path is specific to your system.
# 'chromedriver' is a driver that Selenium uses to open up a new Google Chrome browser.

# chrome_driver_path = '/Users/jonathanbachrach/Documents/Automate/TradedScrape/chromedriver'

# ChromeDriver path
chrome_driver_path = '/usr/local/bin/chromedriver'

# 'Options' is a class in Selenium allowing us to customize & add arguments to Chrome browser.
# The options added here are to prevent any GUI from showing up. 
chrome_options = Options()
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")
chrome_options.binary_location = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"

# Function to extract hashtags
def extract_hashtags(content):
    # the findall() function from the 're' module returns all non-overlapping matches 
    # of pattern in string, as a list of strings.
    hashtags = re.findall(r"#(\w+)", content)
    return ' '.join(['#' + tag for tag in hashtags])

# Function to parse content from each Instagram post.
def parse_content(content):
    soup = BeautifulSoup(content, 'html.parser')

    # Initialize default fields
    data = {'tradedny': None, 'hashtags': extract_hashtags(content)}

    # Normalize keys (label_dict helps match known variants to your Excel headers)
    label_dict = {
        "BROKERS": "BROKER", "NOTE FROM BROKER": "NOTE", "BUYERS": "BUYER", "BUYER'S": "BUYER",
        "SELLERS": "SELLER", "SELLER'S": "SELLER", "BUYERS REP": "BUYER'S REP",
        "SELLERS REP": "SELLER'S REP", "TENANT'S REP": "TENANT REP", "UNIT": "UNITS"
    }

    lines = content.split("\n")

    for line in lines:
        # Skip if it's just hashtags
        if line.startswith("#"):
            continue

        # Only process lines with a colon
        if ":" in line:
            key, value = line.split(":", 1)
            key = key.strip().upper()
            value = BeautifulSoup(value.strip(), 'html.parser').get_text()

            # Normalize using label_dict, fallback to lowercase key
            normalized_key = label_dict.get(key, key).lower()

            data[normalized_key] = value

    return data


# Function to handle the login.
def login(browser, username, password):
    logging.info("Navigating to Instagram login page...")
    browser.get('https://www.instagram.com/accounts/login/')
    time.sleep(5)  # Let the page load

    max_retries = 3

    for attempt in range(max_retries):
        try: 
            # Wait for username input
            logging.info("Waiting for username input field...")
            userElem = WebDriverWait(browser, 30).until(
                EC.presence_of_element_located((By.NAME, 'username'))
            )

            userElem.clear()
            userElem.send_keys(username)

            # Wait for password input
            logging.info("Waiting for password input field...")

            passwordElem = WebDriverWait(browser, 30).until(
                EC.presence_of_element_located((By.NAME, 'password'))
            )

            passwordElem.clear()
            passwordElem.send_keys(password)

            # Submit login form
            passwordElem.send_keys(Keys.RETURN)

            # Wait for login to complete
            logging.info("Waiting for login to complete...")
            WebDriverWait(browser, 30).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "nav"))
            )

            logging.info("Login successful!")

            # Dismiss notification popup if it appears
            try:
                notification_button = WebDriverWait(browser, 10).until(
                    EC.element_to_be_clickable((By.CLASS_NAME, '_a9--'))
                )
                notification_button.click()
                logging.info("Dismissed notification popup.")

            except TimeoutException:
                logging.info("Notification popup not found—skipping.")

            time.sleep(5)
            return # Success: exit the function

        except Exception as e:
            logging.error(f"Login attempt {attempt+1} failed: {type(e).__name__} - {e}")
            logging.exception("Full traceback:")

            time.sleep(10)

    logging.error("Login failed after 3 attempts. Aborting.")
    raise Exception("Login failed.")


# Function to navigate to the profile.
def load_profile(browser, profile_name):
    try:
        """Go to tradedny instagram page."""
        search_element = browser.find_element(By.CSS_SELECTOR, 
            '.x1xgvd2v > div:nth-child(2) > div:nth-child(2) > '
            'span:nth-child(1) > div:nth-child(1) > a:nth-child(1) > '
            'div:nth-child(1) > div:nth-child(2) > div:nth-child(1) > '
            'div:nth-child(1) > span:nth-child(1) > span:nth-child(1)')

        browser.execute_script("arguments[0].scrollIntoView();", search_element)
        time.sleep(2)

        # Click the search element
        browser.execute_script("arguments[0].click();", search_element)

        time.sleep(5)

        # Wait for the search input field and enter the profile name

        type_search = WebDriverWait(browser, 20).until( 
            EC.presence_of_element_located((By.CSS_SELECTOR, 'input.x1lugfcp'))
        )

        type_search.send_keys(PROFILE_NAME)

        # Wait for results to load
        time.sleep(15)

        # Locate the profile link element

        type_click = browser.find_element(
            By.XPATH, 
            "//span[contains(@class, 'x1lliihq') and contains(text(), 'Traded: New York🗽')]" 
        )
    
        # Scroll the profile element into view    
        browser.execute_script("arguments[0].scrollIntoView();", type_click)
        time.sleep(2)

        # Click on the profile link
        type_click.click()
        # Allow time for the page to navigate
        time.sleep(20)
    except (TimeoutException, NoSuchElementException) as e:
        logging.error(f"Loading profile failed: {e}")
        logging.error("Pausing for 5 minutes before retrying...")
        time.sleep(300)
        load_profile(browser, profile_name)  # Retry loading the profile


# Function to scroll down the page to load all posts.
def scroll_down(browser):
    """Scroll to the end of the page to load all posts (add as needed).
    scrollIntoView() method scrolls the specified element into the visible area 
    of the browser window."""

    browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(15)

# Function to click on first post.
def click_on_post(browser): 
    try:
        first_post = WebDriverWait(browser, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 
            "img.x5yr21d.xu96u03.x10l6tqk.x13vifvy.x87ps6o.xh8yej3"))
        )

        # Click the element via JavaScript
        browser.execute_script("arguments[0].click();", first_post)
        time.sleep(4)
    except (TimeoutException, NoSuchElementException) as e:
        logging.error(f"Clicking on the first post failed: {e}")
        logging.error("Pausing for 3 minutes before retrying...")
        time.sleep(180)
        click_on_post(browser)  # Retry clicking the post


# Function to parse a post.
def parse_post(browser, last_text):
    logging.info("Parsing the post content...")

    try:
        # Wait for and extract text from the new robust h1 location
        h1_element = WebDriverWait(browser, 25).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div._a9zr h1"))
        )

        # Extract the text from the h1 element
        new_text = h1_element.text.strip()

        if not new_text or new_text == last_text:
            logging.warning("Duplicate or missing post text.")
            return None, last_text

        logging.info("Post content successfully extracted successfully from <h1>.")
        post_data = parse_content(new_text)
        logging.info(f"Data extracted for post: {post_data}")

        return post_data, new_text

    except Exception as e:
        logging.error(f"Error parsing post: {e}", exc_info=True)
        return None, last_text


# Function to add parsed data to the Excel spreadsheet.
def save_data(browser, data, headers, sheet):
    """The data is a list of values, with each value 
    corresponding to a header in the Excel spreadsheet."""
    # If a header does not have a corresponding value in the data, 
    # use a space " " as a placeholder.
    row = [data.get(header, " ") for header in headers]
    sheet.append(row)


def start_browser(): 
    # Setup the ChromeDriver service with the specified path
    try: 

        service = Service(chrome_driver_path)

        # Use webdriver-manager to download and manage the correct version
        options = Options()
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.binary_location = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"

        logging.info('Initializing ChromeDriver...')

        # Automatically gets the right ChromeDriver for your installed Chrome version
        service = Service(ChromeDriverManager().install())
        browser = webdriver.Chrome(service=service, options=options)

        logging.info('ChromeDriver initialized. Starting Chrome...')
        logging.info('Chrome started successfully.')
        return browser

    except Exception as e: 
        logging.error(f"Failed to initialize ChromeDriver: {e}")
        print(f"Failed to initialize ChromeDriver: {e}")
        exit(1)

    return browser


def read_last_scraped_index():
    """ Check if the file exists and read the last scraped index."""
    if os.path.exists("last_scraped_index.txt"):
        with open("last_scraped_index.txt", "r") as file:
            return int(file.read().strip())
    else:
        return 0


def write_last_scraped_index(index):
    # Write the last scraped index to a file
    with open("last_scraped_index.txt", "w") as file:
        file.write(str(index))


def initialize_browser_and_workbook():
    """Initialize the browser and load or create the Excel workbook."""
    browser = start_browser()
    time.sleep(5)
    browser.get("https://www.instagram.com")
    logging.info('Starting the scraping process...') 

    # Define headers
    headers = ["tradedny", "image", "date", "address", "market", "asset type", "lender", "buyer",
    "renter", "seller", "landlord", "seller's rep", "buyer's rep", "loan amount", "loan type", "tenant", "tenant rep",
    "landlord rep", "broker", "sale price", "asking rent", "sf", "ppsf", "units",
    "ppu", "bsf", "ppbsf", "note", "hashtags"]

    # Load existing workbook if it already exists
    if os.path.exists(OUTPUT_FILE_NAME):
        logging.info('Loading existing Excel workbook...')
        wb = load_workbook(OUTPUT_FILE_NAME)
        sheet = wb.active

    else: 
        # Create a new workbook if one doesn't exist
        logging.info('Initializing the Excel workbook...')
        wb = openpyxl.Workbook()
        sheet = wb.active
        # Insert the headers to the first row.
        sheet.append(headers)
        wb.save(OUTPUT_FILE_NAME)  # Force immediate save to disk

    # Debugging: Log the outputs
    if browser and wb and sheet and headers: 
        logging.info('Browser, workbook, sheet, and headers initialized successfully.')
    else: 
        logging.error('One or more of the initialization components is None.')

    return browser, wb, sheet, headers 


def scrape_instagram(browser, wb, sheet, headers, last_text, num_posts=10):
    """Perform the Instagram scraping process."""
    logging.info(f"Starting the scraping process...")

    profile_url = "https://www.instagram.com/tradedny/"
    browser.get(profile_url)
    time.sleep(5)

    # Click on the first post
    try:
        logging.info("Waiting for the first post to become clickable...")
        first_post = WebDriverWait(browser, 20).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, "div._aagw"))
        )
        browser.execute_script("arguments[0].scrollIntoView(true);", first_post)
        time.sleep(2)
        first_post.click()
        logging.info("Clicked on the first post.")
        time.sleep(5)
    except Exception as e:
        logging.error(f"Failed to open the first post: {e}", exc_info=True)
        return

    consecutive_errors = 0

    for i in range(num_posts):
        logging.info(f"Scraping post {i + 1} of {num_posts}...")
        try:
            # Call parse_post() to get values for post_data and new_text
            post_data, new_text = parse_post(browser, last_text)

            if not post_data:
                logging.warning("No data extracted from post. Skipping...")
                consecutive_errors += 1
                if consecutive_errors >= 3:
                    raise Exception("Too many consecutive errors. Exiting the program.")
                continue

            last_text = new_text

            logging.info(f"Data extracted for post {i + 1}: {post_data}")
            logging.info("Saving post data to Excel...")

            row = [post_data.get(header.lower(), "") for header in headers]
            logging.info(f"Appending row: {row}")
            sheet.append(row)
            wb.save(OUTPUT_FILE_NAME)

            consecutive_errors = 0

            # Save workbook every 10 posts
            if (i + 1) % 10 == 0:
                wb.save(OUTPUT_FILE_NAME)
                logging.info("Saved Excel workbook after 10 posts.")

            # Click the "Next" arrow to go to the next post
            try:
                next_button = WebDriverWait(browser, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, "div._aaqg button._abl-"))
                )
                next_button.click()
                logging.info("Clicked next post.")
                time.sleep(5)
            except TimeoutException:
                logging.warning("Next button not found. Attempting to continue...")
                break 

        except Exception as e:
            logging.error(f"Unexpected error scraping post {i + 1}: {e}", exc_info=True)
            consecutive_errors += 1
            if consecutive_errors >= 3:
                logging.error("Too many consecutive errors scraping.")
                raise Exception("Too many consecutive errors. Exiting the program.")
            continue

    # Final save after all posts are processed
    wb.save(OUTPUT_FILE_NAME)
    logging.info("Saved Excel workbook after scraping all posts.")


def save_workbook_and_cleanup(wb, browser):
    """Save the workbook and clean up resources."""
    if wb: 
        logging.info('Saving the Excel workbook...')
        wb.save(OUTPUT_FILE_NAME)

    if browser: 
        logging.info('Closing the browser...')
        browser.quit()

    if os.path.exists("last_scraped_index.txt"):
        os.remove("last_scraped_index.txt")
        logging.info('Deleted last_scraped_index.txt to start fresh next time.')


def main():
    logging.info('Starting the scraping process...')

    # Initialize browser and workbook variables
    browser = None 
    wb = None
    last_text = ""

    try:
        browser, wb, sheet, headers = initialize_browser_and_workbook()
        if not all([browser, wb, sheet, headers]):
            logging.error("Initialization failed: One or more components are None.")
            return 

        # Log into Instagram before scraping
        login(browser, USERNAME, PASSWORD)

        while True:
            try: 
                # Perform the scraping process
                scrape_instagram(browser, wb, sheet, headers, last_text, num_posts=NUM_POSTS_TO_SCRAPE)
                # Exit the loop if scraping completes successfully
                break

            except (WebDriverException, InvalidSessionIdException) as e:
                logging.error(f"Browset crashed: {e}")
                if browser: 
                    try:
                        browser.quit()
                    except:
                        pass
                logging.info("Restarting browser and reinitializing workbook...") 
                # Slight delay before attempting to restart
                browser, wb, sheet, headers = initialize_browser_and_workbook()
                login(browser, USERNAME, PASSWORD)

            except Exception as e:
                logging.error(f"Unexpected error in main loop: {e}")
                # Exit the loop on other unexpected errors
                break 


    finally: 
        save_workbook_and_cleanup(wb, browser)


# Call the main function.
if __name__ == "__main__":
    main()

